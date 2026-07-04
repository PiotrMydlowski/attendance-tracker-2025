"""
Created 2026
Author: Piotr M

Loosely based on a web tutorial.
This code checks for the amount of absences in spreadsheet file
and sends notification to matching email in case of too many.
"""
import openpyxl
import smtplib
import ssl

from credentials import *
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText


def readFile():
    """Initialization function."""
    book = openpyxl.load_workbook('attendance.xlsx') #opening workbook
    sheet = book['Sheet 1'] # Choose the sheet
    r = sheet.max_row # counting number of rows / students
    c = sheet.max_column # counting number of columns / subjects
    return r, c, sheet


def countStudentAbsences(row, columns, sheet):
    """Summarizes absences"""
    a = 0
    for lesson in range(2, columns+1):
        presence = sheet.cell(row=row, column=lesson).value
        if presence=="Absent":
            a = a +1
    return a


def sendMail(to_id, absences):
    """Sends mail to students with their absences """
    msg = "You have been already absent " + str(absences) + " times."
    message = MIMEMultipart()
    message['Subject'] = 'Attendance report'
    message['From'] = login
    message['To'] = to_id
    message.attach(MIMEText(msg, 'plain'))
    content = message.as_string()
    print('Mail prepared.')

    s = smtplib.SMTP_SSL(hostName, portNumber, timeout=120)
    context = ssl.create_default_context()
    s.ehlo()
    #s.starttls(context=context)
    s.ehlo()
    s.login(login, secret)
    s.sendmail(from_addr=login, to_addrs=to_id, msg=content)
    s.quit()
    print('Mail sent.')
    return 0


def main():
    """Main function."""
    (rows, columns, sheet) = readFile()
    for row in range (2, rows+1):
        to_id = sheet.cell(row=row, column=1).value
        absences = countStudentAbsences(row, columns, sheet)
        if absences > 6:
            sendMail(to_id, absences)

    print("Code execution finished.")


if __name__ == "__main__":
    main()